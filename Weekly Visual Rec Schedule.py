import requests
from ics import Calendar
from datetime import datetime, timedelta
import re, ssl, pytz, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# --- iCal Feed ---
ical_url = "http://tmsdln.com/19hyx"
local_tz = pytz.timezone("America/New_York")
cutoff_date = datetime(2025, 9, 6).date()

# --- Color Map ---
color_map = {
    "Blue": "#4996D1",
    "Red": "#E88989",
    "Green": "#429964",
    "Orange": "#FCB03A",
    "Berry": "#E8DAEF",
    "Gray": "#F2F3F4"
}

# --- Helpers ---
def parse_team(raw_team):
    match = re.match(r"(.+?)\s*\((.*?)\s*-\s*(.*?)\)", raw_team)
    if match:
        name = match.group(1).strip()
        coach = match.group(2).strip()
        color = match.group(3).strip()
        return f"{name} ({coach})", color
    return raw_team.strip(), "Gray"

def extract_group(team_name):
    match = re.search(r"(\d+(?:/\d+)*\s+(Boys|Girls))", team_name)
    if match:
        return match.group(1)
    if "Kindergarten" in team_name:
        return "Kindergarten"
    return ""

def extract_division(description):
    match = re.search(r"(\d+(?:/\d+)*\s+(Boys|Girls))", description or "")
    if match:
        return match.group(1)
    if "Kindergarten" in description:
        return "Kindergarten"
    return ""

def format_field(raw_field):
    if "H-SuSS" in raw_field:
        return "Snack Shack Area"
    if not raw_field or len(raw_field) < 5:
        return raw_field
    trimmed = raw_field[4:].split(",")[0].strip()
    match = re.match(r"([A-Z]*)(\d+)([A-Z]*)", trimmed)
    if match:
        _, number, suffix = match.groups()
        return f"Field {number}{suffix}" if suffix else f"Field {number}"
    return f"Field {trimmed}"

def field_sort_key(field_label):
    match = re.match(r"Field\s+(\d+)([A-Z]?)", field_label)
    if match:
        number = int(match.group(1))
        suffix = match.group(2)
        return (number, suffix)
    return (999, "")

# --- Load Calendar ---
ssl._create_default_https_context = ssl._create_unverified_context
calendar = Calendar(requests.get(ical_url).text)

# --- Group Events ---
games_by_date = defaultdict(list)

for event in calendar.events:
    local_start = event.begin.datetime.astimezone(local_tz)
    game_date = local_start.date()
    if game_date <= cutoff_date:
        continue

    time_label = local_start.strftime("%I:%M %p").lstrip("0")
    sort_key = local_start.strftime("%H:%M")
    name = event.name
    location = event.location or ""
    description = event.description or ""

    if "Practice" in name:
        continue

    if any(k in name for k in ["3/4", "5/6", "7/8"]) and "vs." in name:
        continue

    if "vs." in name:
        team1_raw, team2_raw = name.split("vs.")
        team1, color1 = parse_team(team1_raw.strip())
        team2, color2 = parse_team(team2_raw.strip())
        group = extract_group(team1_raw.strip()) or extract_group(team2_raw.strip())
    else:
        continue  # Skip non-match events

    field = format_field(location)
    division = extract_division(description)

    games_by_date[game_date].append(
        [sort_key, time_label, field, team1, color1, team2, color2, group, division]
    )

# --- Excel Output (optional, saved to repo root) ---
excel_file = "non_core_games.xlsx"
wb = Workbook()
wb.remove(wb.active)

for game_date in sorted(games_by_date.keys()):
    ws = wb.create_sheet(title=game_date.strftime("%Y-%m-%d"))
    ws.append(["Time", "Field", "Team 1", "Team 2", "Group", "Division"])
    sorted_games = sorted(games_by_date[game_date], key=lambda x: (x[0], field_sort_key(x[2])))
    row_index = 2
    for _, time_label, field, team1, color1, team2, color2, group, division in sorted_games:
        if division == "":
            continue
        ws.append([time_label, field, team1, team2, group, division])
        fill1 = PatternFill(start_color=color_map.get(color1)[1:], end_color=color_map.get(color1)[1:], fill_type="solid")
        fill2 = PatternFill(start_color=color_map.get(color2)[1:], end_color=color_map.get(color2)[1:], fill_type="solid")
        ws[f"C{row_index}"].fill = fill1
        ws[f"D{row_index}"].fill = fill2
        row_index += 1

wb.save(excel_file)

# --- HTML Output ---
def write_html(filename, game_date, games):
    with open(filename, "w", encoding="utf-8") as f:
        f.write("<html><head><style>\n")
        f.write("""
        body { font-family: sans-serif; padding: 20px; }
        .time-group { margin-bottom: 40px; }
        .time-group h2 { margin-bottom: 10px; }
        .match-row { display: flex; flex-wrap: wrap; gap: 20px; }
        .match-wrapper { margin-bottom: 50px; position: relative; }
        .division-label {
            margin-bottom: 5px;
            font-size: 0.85em;
            color: #888;
            text-align: center;
        }
        .group-label {
            margin-bottom: 5px;
            font-size: 0.8em;
            color: #444;
            text-align: center;
        }
        .match-block {
            display: flex;
            width: 300px;
            height: 80px;
            border: 2px solid #ccc;
            font-weight: bold;
            color: #333;
        }
        .team-left, .team-right {
            flex: 1;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 5px;
        }
        .team-left { border-right: 1px solid #aaa; }
        .field-label {
            position: absolute;
            bottom: -20px;
            left: 50%;
            transform: translateX(-50%);
            font-size: 0.9em;
            color: #666;
        }
        """)
        f.write("</style></head><body>\n")
        f.write(f"<h1>{game_date.strftime('%A, %B %d, %Y')}</h1>\n")

        time_groups = defaultdict(list)
        for row in games:
            time_groups[row[1]].append(row)

        ordered_times = sorted(time_groups.keys(), key=lambda t: datetime.strptime(t, "%I:%M %p"))
        for time in ordered_times:
            f.write(f"<div class='time-group'><h2>{time}</h2><div class='match-row'>\n")
            for _, _, field, team1, color1, team2, color2, group, division in sorted(time_groups[time], key=lambda x: field_sort_key(x[2])):
                if division == "":
                    continue
                f.write("<div class='match-wrapper'>\n")
                f.write(f"<div class='division-label'>{division}</div>\n")
                f.write(f"<div class='group-label'>{group}</div>\n")
                f.write(f"""
                <div class="match-block">
                    <div class="team-left" style="background-color:{color_map.get(color1)};">{team1}</div>
                    <div class="team-right" style="background-color:{color_map.get(color2)};">{team2}</div>
                    <div class="field-label">{field}</div>
                </div>
                """)
                f.write("</div>\n")
            f.write("</div></div>\n")

        f.write("</body></html>")

# Write public-facing HTML for next Saturday
today = datetime.now(local_tz).date()
days_until_saturday = (5 - today.weekday()) % 7
next_saturday = today + timedelta(days=days_until_saturday)

if next_saturday in games_by_date:
    html_week_file = "index.html"
    write_html(html_week_file, next_saturday, games_by_date[next_saturday])
    print(f"✅ This week's layout saved to: {html_week_file}")
else:
    print("⚠️ No games found for the upcoming Saturday.")
