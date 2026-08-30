import requests
from ics import Calendar
from datetime import datetime, timedelta
import re, ssl, pytz
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import random

# --- ICS Feed ---
ical_url = "https://calendar.google.com/calendar/ical/6bl9ubrc8vssoqi0jm1l7ljpc05ngqrt%40import.calendar.google.com/public/basic.ics"

local_tz = pytz.timezone("America/New_York")
today = datetime.now(local_tz).date()

# --- Helpers ---
def parse_team(raw_team):
    match = re.match(r"(.+?)\s*`\((.*?)\s*-\s*(.*?)\)`", raw_team)
    if match:
        name = match.group(1).strip()
        coach = match.group(2).strip()
        color = match.group(3).strip()
        return f"{name} ({coach})", color
    return raw_team.strip(), "Gray"

def extract_group(team_name):
    match = re.search(r"(\d+(?:/\d+)*\s+(Boys|Girls))", team_name)
    if match:
        return match.group(1)
    if "Kindergarten" in team_name:
        return "Kindergarten"
    return ""

def extract_division(description):
    match = re.search(r"(\d+(?:/\d+)*\s+(Boys|Girls))", description or "")
    if match:
        return match.group(1)
    if "Kindergarten" in description:
        return "Kindergarten"
    return ""

def format_field(raw_field):
    if "H-SuSS" in raw_field:
        return "Snack Shack Area"
    if not raw_field or len(raw_field) < 5:
        return raw_field
    trimmed = raw_field[4:].split(",")[0].strip()
    match = re.match(r"([A-Z]*)(\d+)([A-Z]*)", trimmed)
    if match:
        _, number, suffix = match.groups()
        return f"Field {number}{suffix}" if suffix else f"Field {number}"
    return f"Field {trimmed}"

def field_sort_key(field_label):
    match = re.match(r"Field\s+(\d+)([A-Z]?)", field_label)
    if match:
        return (int(match.group(1)), match.group(2))
    return (999, "")

# --- Auto Color Detection ---
def get_color_map_from_schedule(future_games):
    known_colors = {
        "Blue": "#4996D1",
        "Red": "#E88989",
        "Green": "#429964",
        "Orange": "#FCB03A",
        "Berry": "#E8DAEF",
        "Gray": "#F2F3F4",
        "Black": "#000000",
        "White": "#FFFFFF",
        "Yellow": "#FFEB3B",
        "Purple": "#9B59B6",
        "Maroon": "#800000",
        "Teal": "#008080",
        "Pink": "#FFC0CB",
        "Gold": "#FFD700",
        "Silver": "#C0C0C0",
        "Navy": "#001F3F",
        "Royal Blue": "#4169E1",
        "Light Blue": "#ADD8E6",
        "Dark Green": "#006400",
    }

    auto_map = {}

    for date, games in future_games.items():
        for g in games:
            color1 = g[3]
            color2 = g[5]

            for c in (color1, color2):
                if c not in auto_map:
                    auto_map[c] = known_colors.get(
                        c,
                        "#{:06x}".format(random.randint(0x444444, 0xDDDDDD))
                    )

    return auto_map

# --- Load ICS ---
ssl._create_default_https_context = ssl._create_unverified_context
ics_text = requests.get(ical_url).text

if "<!DOCTYPE html>" in ics_text[:200]:
    raise Exception("ICS feed returned HTML instead of ICS.")

calendar = Calendar(ics_text)

# --- Extract all future REC games ---
future_games = defaultdict(list)

for event in calendar.events:
    local_start = event.begin.datetime.astimezone(local_tz)
    game_date = local_start.date()

    if game_date < today:
        continue

    name = event.name
    location = event.location or ""
    description = event.description or ""

    if "Practice" in name or "vs." not in name:
        continue

    time_label = local_start.strftime("%I:%M %p").lstrip("0")
    team1_raw, team2_raw = name.split("vs.")
    team1, color1 = parse_team(team1_raw.strip())
    team2, color2 = parse_team(team2_raw.strip())
    field = format_field(location)
    group = extract_group(team1_raw.strip()) or extract_group(team2_raw.strip())
    division = extract_division(description)

    # ⭐ Skip Travel games
    if "Travel" in division:
        continue

    future_games[game_date].append(
        [time_label, field, team1, color1, team2, color2, group, division]
    )

# --- Auto Color Map ---
color_map = get_color_map_from_schedule(future_games)

def safe_color(c):
    return color_map.get(c, "#DDDDDD")[1:]  # remove '#'

# --- Determine upcoming Saturday ---
next_saturday = today + timedelta((5 - today.weekday()) % 7)
games_this_sat = future_games.get(next_saturday, [])

# --- Determine next REC game day ---
next_game_date = None
for d in sorted(future_games.keys()):
    if future_games[d]:
        next_game_date = d
        break

# --- Excel Output ---
excel_file = "non_core_games.xlsx"
wb = Workbook()
ws_default = wb.active
ws_default.title = "Placeholder"

for game_date in sorted(future_games.keys()):
    games = future_games[game_date]
    if not games:
        continue

    ws = wb.create_sheet(title=game_date.strftime("%Y-%m-%d"))
    ws.append(["Time", "Field", "Team 1", "Team 2", "Group", "Division"])

    row_index = 2
    for time_label, field, team1, color1, team2, color2, group, division in sorted(
        games,
        key=lambda x: (datetime.strptime(x[0], "%I:%M %p"), field_sort_key(x[1]))
    ):
        ws.append([time_label, field, team1, team2, group, division])
        ws[f"C{row_index}"].fill = PatternFill(start_color=safe_color(color1), end_color=safe_color(color1), fill_type="solid")
        ws[f"D{row_index}"].fill = PatternFill(start_color=safe_color(color2), end_color=safe_color(color2), fill_type="solid")
        row_index += 1

if len(wb.sheetnames) > 1:
    del wb["Placeholder"]
else:
    ws_default.append(["No REC games found"])

wb.save(excel_file)

# --- HTML Output ---
html_file = "index.html"

with open(html_file, "w", encoding="utf-8") as f:
    f.write("<html><head><style>\n")
    f.write("""
        body { font-family: sans-serif; padding: 20px; }
        .match-row { display: grid; grid-template-columns: repeat(auto-fill, minmax(240px, 1fr)); gap: 12px; }
        .match-wrapper { border: 1px solid #000; padding: 10px; margin-bottom: 20px; }
        .team-left, .team-right { font-weight: bold; padding: 4px; }
        .division-label { font-size: 0.85em; margin-top: 4px; }
    """)
    f.write("</style></head><body>\n")

    # Small message if no Rec games this Saturday
    if not games_this_sat:
        f.write(f"""
        <p style="color:#666; font-style:italic; font-size:0.85em;">
            No REC games scheduled for Saturday, {next_saturday.strftime('%B %d')}
        </p>
        """)

    # Next REC game day heading
    if next_game_date:
        f.write(f"<h1>Next REC game day: {next_game_date.strftime('%A, %B %d')}</h1>\n")
    else:
        f.write("<h1>No upcoming REC games found.</h1>")
        f.write("</body></html>")
        exit(0)

    # Render next REC game day grid
    games = future_games[next_game_date]
    time_groups = defaultdict(list)
    for row in games:
        time_groups[row[0]].append(row)

    for time in sorted(time_groups.keys(), key=lambda t: datetime.strptime(t, "%I:%M %p")):
        f.write(f"<h2>{time}</h2><div class='match-row'>\n")

        for time_label, field, team1, color1, team2, color2, group, division in sorted(
            time_groups[time],
            key=lambda x: field_sort_key(x[1])
        ):
            f.write("<div class='match-wrapper'>\n")
            f.write(f"<div class='team-left' style='background-color:{color_map.get(color1)}'>{team1}</div>\n")
            f.write(f"<div class='team-right' style='background-color:{color_map.get(color2)}'>{team2}</div>\n")
            f.write(f"<div class='division-label'>{division} — {group}</div>\n")
            f.write(f"<div style='font-size:0.75em; margin-top:4px;'>Field: {field}</div>\n")
            f.write("</div>\n")

        f.write("</div>\n")

    f.write("</body></html>")

print(f"HTML saved to: {html_file}")
print(f"Excel saved to: {excel_file}")
